# IMPORTATIONS
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException as FileExc
# from openpyxl.chart import LineChart, Reference

from decimal import *
from scipy.special import softmax
from scipy.stats import entropy, linregress
import numpy as np

# ___________________________________________________________
# -----------------------------------------------------------
#   LOADING ORIGINAL DATA FILE 
# -----------------------------------------------------------

# Load the file. Turn read_only to True if necessary
# Example is "geophysics_clean.xlsx" (with NO quotation marks)
data_name = input("Input the data file's name: ")

def load_file(inp):
    try:
        orig_wb = load_workbook(filename = data_name,
                                read_only = True)
        valid_name = True
    except FileExc:
        valid_name = False
    except PermissionError:
        valid_name = "permission"
    except:
        valid_name = False
    return valid_name

while True:
    load_valid = load_file(data_name)
    if type(load_valid) == bool:
        if load_valid:
            break
        else:
            data_name = input("Wrong file name. Try again: ")
    else:
        data_name = input("Close file and input again: ")

# Actually load
orig_wb = load_workbook(filename = data_name, read_only = True)

# Workbook ~ spreadsheet ~ wb
# Worksheet ~ sheet ~ 'sheet_name'

# Working with same-named sheet e.g. "geophysics_clean"
datasheet = orig_wb[data_name[:-5]]


# ___________________________________________________________
# -----------------------------------------------------------
#   NEW FILE + YEAR RANGE (START, YEARS) + START_ROW
# -----------------------------------------------------------

# FUNCTION: check that input is integer
def check_integer(inp):
    try:
        inp = int(inp)
        inp_is_int = True
    except ValueError:
        inp_is_int = False
    return inp_is_int

# New file with (1) Start year = START ; (2) No. of years/sheets = YEARS
wb = Workbook()

START = input("Input start year (inclusive): ")
while True:
    if check_integer(START):
        START = int(START)
        break
    else:
        START = input("Not integer. Input integer: ")

YEARS = 0
TERMIN = input("Input end year (inclusive): ")
while True:
    if check_integer(TERMIN):
        YEARS = int(TERMIN) - START + 1
        break
    else:
        TERMIN = input("Not integer. Input integer: ")


# Create a sheet for each year
for i in range(YEARS):
    ws_temp = wb.create_sheet(str(START+i))


# New workbook xlsx file name = old one + "_processed.xlsx"
FILE_NAME = data_name[:-5] + "_processed"

# Check whether open + Save new file
try:
    wb.save(FILE_NAME+".xlsx")
except:
    temp = input("Close the file and hit enter.")
    wb.save(FILE_NAME+".xlsx")

wb.save(FILE_NAME+".xlsx")

# ------------------- FILE IS SAVED HERE ---------------------

# Identify start row
START_ROW = 1

def start_row_id(start_year):
    for i in range(1, 195):
        if datasheet['B'+str(i)].value == int(start_year):
            start_row = i
            break
    return i

START_ROW = start_row_id(START)

# Function to process a single cell
def process_cell_data(index, wksht):
    '''
    index str, wksht str.
    Extract one cell from datasheet > Deconstruct > Write in new
    '''
    cell_val = datasheet[index].value
    # Remove left end "[[('" and right end ")]]"
    cell_val = cell_val[4:-3]
    # Split into "stem', #" format
    cell_li = cell_val.split("), ('")

    # Report on progress
    print(index, end=' ')

    # Iterate
    for i in range(len(cell_li)):
        temp_li = cell_li[i].split("', ")
        
        # Add stem to one column (A)
        wksht['A'+str(i+1)] = temp_li[0]
        # Add absolute frequency to other column (B)
        wksht['B'+str(i+1)] = int(temp_li[1])
        # Add relative frequency (/NoD) to third column (C)
        wksht['C'+str(i+1)] = int(temp_li[1]) / \
                              datasheet["D"+index[1:]].value

# Back to Main code: column R for top 25%
for i in range(YEARS):
    process_cell_data("R"+str(i+START_ROW), wb[str(START+i)])

# Save new file
wb.save(FILE_NAME+".xlsx")

# ------------------- FILE IS SAVED HERE ---------------------


# ___________________________________________________________
# -----------------------------------------------------------
#   Calculate Shannon ENTROPY (base e)
# -----------------------------------------------------------

# Define Function for rounding 4/5 (half up)
def rounding(number, precision):
    '''
    number float, precision float
    '''
    ret = float(
        Decimal(str(number)).quantize(
            Decimal(str(precision)), rounding=ROUND_HALF_UP
            )
        )
    return ret


# Create new list for the entropies
entropy_li = []

def entropy_list_append(year, entr_li, wkbook):
    col = list(wkbook[str(year)]['C'])
    col_vals = np.array([i.value for i in col])
    entr_rounded = rounding(entropy(col_vals), 0.001)
    entr_li.append(entr_rounded)

for d in range(YEARS):
    entropy_list_append(START+d, entropy_li, wb)

# Create new worksheet for entropy
entr_ws = wb.create_sheet("entropy", 0)

entr_ws['A2'] = "Entropy_0"

# Add the new entropies to the new worksheet
for d in range(YEARS):
    entr_ws.cell(row = 1, column = d+2, value = START+d)
    entr_ws.cell(row = 2, column = d+2, value = entropy_li[d])


wb.save(FILE_NAME+'.xlsx')
# ------------------- FILE IS SAVED HERE ---------------------

