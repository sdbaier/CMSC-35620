# IMPORTATIONS
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException as FileExc

from decimal import *
from scipy.special import softmax
from scipy.stats import entropy, linregress
import numpy as np

# ___________________________________________________________
# -----------------------------------------------------------
#   LOADING ORIGINAL DATA FILE 
# -----------------------------------------------------------

# Load the file. Need to add analysis, so read_only = False
# Example is "geophysics_clean.xlsx" (with NO quotation marks)
data_name = input("Input the data file's name: ")

def load_file(inp):
    try:
        orig_wb = load_workbook(filename = data_name,
                                read_only = False)
        valid_name = True
    except FileExc:
        valid_name = False
    except PermissionError:
        valid_name = "permission"
    except:
        valid_name = False
    return valid_name

while True:
    load_valid = load_file(data_name)
    if type(load_valid) == bool:
        if load_valid:
            break
        else:
            data_name = input("Wrong file name. Try again: ")
    else:
        data_name = input("Close file and input again: ")

# Actually load
orig_wb = load_workbook(filename = data_name, read_only = False)


# ___________________________________________________________
# -----------------------------------------------------------
#   Linear REGRESSION + Print results in the same worksheet
# -----------------------------------------------------------

# Initialize year values
START = 2000
YEARS = 21

# Load original data i.e. datasheet
datasheet = orig_wb['HF25']

# Create new worksheet i.e. ws
ws = orig_wb.create_sheet('linear_reg', 0)

# Do linear regression
def do_lin_regress(Y):
    X = [i for i in range(START, START+YEARS)]
    # Y is the list of entropy or quotient values
    slope, intercept, r_value, p_value, SE = linregress(X,Y)
    return tuple((slope, intercept, r_value, p_value, SE))

# ORIGINAL DATA AREAS:
# Base e entropy -      'B2':'V8' where B=2, V=22
# Normalized entropy -  'B29':'V35'
# Quotient -            'B57':'V63'
# Metric entropy -      'B67':'V73'

# DO LINEAR REGRESSION
def iter_lin_regress(linreg_list, ROW):
    for i in range(7):
        Y_list = list(list(datasheet.iter_rows(i+ROW, i+ROW, 2, 22,
                                               values_only=True))[0])
        linreg_list.append(do_lin_regress(Y_list))
    return linreg_list

# (1) base e entropy from 2
linreg_0 = []
iter_lin_regress(linreg_0, 2)

# (2) normalized entropy from 29
linreg_Norm = []
iter_lin_regress(linreg_Norm, 29)

# (3) quotient from 57
linreg_Quot = []
iter_lin_regress(linreg_Quot, 57)

# (4) metric entropy from 67
linreg_Metr = []
iter_lin_regress(linreg_Metr, 67)

# Print out the values in the worksheet
var_names = ('slope', 'B', 'r', 'p', 'sl. SE')

def out_linreg(ROW, linreg_li):
    for index in range(5):
        ws.cell(row = ROW-1, column = index+2, value = var_names[index])
    for i in range(7):
        for j in range(5):
            ws.cell(row = ROW+i, column = j+2, value = linreg_li[i][j])

# Actual printout
out_linreg(2, linreg_0)
out_linreg(11, linreg_Norm)
out_linreg(20, linreg_Quot)
out_linreg(29, linreg_Metr)

orig_wb.save(data_name[:-5]+"_linreg.xlsx")
